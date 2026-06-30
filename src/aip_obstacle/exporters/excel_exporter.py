"""Excel 导出器：按解析顺序写出完整障碍物表。"""

from __future__ import annotations

from pathlib import Path
from typing import List

from aip_obstacle.models import Obstacle

_TOTAL_COLS = 7  # 原始编号+名称+方位+距离+海拔+场压+人工修改标记


def export_xlsx(obstacles: List[Obstacle], out_path: str | Path) -> None:
    """把障碍物列表写成单表 Excel，保留输入顺序和 AIP 原始编号。

    说明：
    - 不再按距离拆成左右栏，避免覆盖 PDF 原始表头分区。
    - 不再按距离排序，保持解析/入库顺序。
    - 序号列使用 AIP 原始 obstacle_id，而不是重新从 1 编号。
    依赖 openpyxl，未安装时抛出 ImportError。
    """
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Alignment, Border, Font, Side
        from openpyxl.utils import get_column_letter
    except ImportError:
        raise ImportError("请先安装 openpyxl：pip install openpyxl")

    nrows = len(obstacles) + 2  # +2 行表头

    wb = Workbook()
    ws = wb.active
    ws.title = "机场细则障碍物"

    # --- 样式 ---
    thin_border = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"), bottom=Side(style="thin"),
    )
    header_font = Font(bold=True, size=10)
    title_font = Font(bold=True, size=11)
    wrap_align = Alignment(horizontal="center", vertical="center", wrap_text=True)

    def _border_row(r: int):
        for c in range(1, _TOTAL_COLS + 1):
            ws.cell(row=r, column=c).border = thin_border

    # --- 第 1 行：章节标题 ---
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=_TOTAL_COLS)
    c = ws.cell(row=1, column=1, value="主要障碍物")
    c.font = title_font
    c.alignment = Alignment(horizontal="center", vertical="center")

    # --- 第 2 行：列标题 ---
    headers = [
        "原始编号",
        "障碍物名称",
        "磁方位（度）",
        "相对跑道基准点距离（m）",
        "海拔高度（m）",
        "场压高度（m）",
        "是否人工修改",
    ]
    for ci, h in enumerate(headers, start=1):
        cell = ws.cell(row=2, column=ci, value=h)
        cell.font = header_font
        cell.alignment = wrap_align

    # --- 数据行 ---
    for idx, obs in enumerate(obstacles):
        row = idx + 3
        ws.cell(row=row, column=1, value=obs.obstacle_id)
        ws.cell(row=row, column=2, value=obs.name)
        if obs.mag_bearing_deg is not None:
            ws.cell(row=row, column=3, value=int(obs.mag_bearing_deg))
        if obs.distance_m is not None:
            ws.cell(row=row, column=4, value=int(obs.distance_m))
        if obs.elevation_m is not None:
            c = ws.cell(row=row, column=5, value=round(obs.elevation_m, 1))
            c.number_format = "0.0"
        # col 6 = 场压高度，留空
        ws.cell(row=row, column=7, value="是" if obs.is_user_modified else "否")

    # --- 边框 ---
    for r in range(1, nrows + 1):
        _border_row(r)

    # --- 列宽 ---
    width_map = {1: 6, 2: 28, 3: 12, 4: 14, 5: 12, 6: 10, 7: 12}
    for ci, w in width_map.items():
        ws.column_dimensions[get_column_letter(ci)].width = w

    # --- 行高 ---
    ws.row_dimensions[1].height = 22
    ws.row_dimensions[2].height = 36
    for r in range(3, nrows + 1):
        ws.row_dimensions[r].height = 28

    # --- 写入 ---
    p = Path(out_path)
    p.parent.mkdir(parents=True, exist_ok=True)
    wb.save(str(p))
