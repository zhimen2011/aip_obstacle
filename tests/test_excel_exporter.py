"""测试：Excel 导出保持原始顺序和原始编号。"""

from openpyxl import load_workbook

from aip_obstacle.exporters.excel_exporter import export_xlsx
from aip_obstacle.models import Obstacle


def test_export_xlsx_keeps_original_order_and_ids(tmp_path):
    obstacles = [
        Obstacle(
            airport_icao="ZGSZ",
            obstacle_id="001",
            name="望牛亭",
            mag_bearing_deg=4.0,
            distance_m=4187.0,
            elevation_m=113.7,
        ),
        Obstacle(
            airport_icao="ZGSZ",
            obstacle_id="058",
            name="T3新塔台",
            mag_bearing_deg=252.0,
            distance_m=753.0,
            elevation_m=94.0,
            is_user_modified=True,
        ),
        Obstacle(
            airport_icao="ZGSZ",
            obstacle_id="081",
            name="莲花山",
            mag_bearing_deg=3.0,
            distance_m=23239.0,
            elevation_m=511.0,
        ),
    ]

    out_path = tmp_path / "obstacles.xlsx"
    export_xlsx(obstacles, out_path)

    ws = load_workbook(out_path).active

    assert ws.cell(row=1, column=1).value == "主要障碍物"
    assert ws.cell(row=2, column=1).value == "原始编号"
    assert ws.cell(row=2, column=2).value == "障碍物名称"
    assert ws.cell(row=2, column=7).value == "是否人工修改"

    assert ws.cell(row=3, column=1).value == "001"
    assert ws.cell(row=3, column=2).value == "望牛亭"
    assert ws.cell(row=3, column=7).value == "否"
    assert ws.cell(row=4, column=1).value == "058"
    assert ws.cell(row=4, column=2).value == "T3新塔台"
    assert ws.cell(row=4, column=7).value == "是"
    assert ws.cell(row=5, column=1).value == "081"
    assert ws.cell(row=5, column=2).value == "莲花山"
