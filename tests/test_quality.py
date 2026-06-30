from __future__ import annotations

from pathlib import Path

import pytest
from openpyxl import load_workbook

from aip_obstacle.models import Obstacle, ParseFailure
from aip_obstacle.pipeline import _build_parse_result, parse_file
from aip_obstacle.services.quality import (
    CONFIDENCE_HIGH_RISK,
    CONFIDENCE_OK,
    append_missing_id_failures,
    apply_quality_scores,
    merge_table_and_text_results,
)


def _obs(
    obstacle_id: str,
    mag_bearing_deg: float | None,
    distance_m: float | None,
    elevation_m: float | None = 1000.0,
) -> Obstacle:
    return Obstacle(
        airport_icao="ZLLL",
        obstacle_id=obstacle_id,
        name="test",
        mag_bearing_deg=mag_bearing_deg,
        distance_m=distance_m,
        elevation_m=elevation_m,
        source_page=1,
        raw_text=f"{obstacle_id} raw",
    )


def test_missing_latlon_does_not_lower_confidence():
    scored = apply_quality_scores([_obs("001", 10.0, 1000.0)])

    assert scored[0].latitude is None
    assert scored[0].longitude is None
    assert scored[0].confidence_score == CONFIDENCE_OK


def test_missing_elevation_is_high_risk():
    scored = apply_quality_scores([_obs("001", 10.0, 1000.0, elevation_m=None)])

    assert scored[0].confidence_score == CONFIDENCE_HIGH_RISK


def test_tiny_distance_is_high_risk():
    scored = apply_quality_scores([_obs("001", 10.0, 36.0)])

    assert scored[0].confidence_score == CONFIDENCE_HIGH_RISK


def test_bearing_order_anomaly_is_high_risk():
    scored = apply_quality_scores([
        _obs("059", 35.0, 20000.0),
        _obs("060", 65.0, 21000.0),
        _obs("061", 50.0, 22000.0),
    ])

    assert scored[0].confidence_score == CONFIDENCE_OK
    assert scored[1].confidence_score == CONFIDENCE_HIGH_RISK
    assert scored[2].confidence_score == CONFIDENCE_OK


def test_table_failure_is_filled_from_text_fallback():
    failure = ParseFailure(
        airport_icao="ZLLL",
        source_page=6,
        raw_text="山\n060 | 山 | | N364859 | | 2484",
        reason="编号 060：缺少方位/距离",
    )
    text = _obs("060", 4.0, 33630.0, elevation_m=2484.0)

    obstacles, failures = merge_table_and_text_results([], [failure], [text])

    assert failures == []
    assert len(obstacles) == 1
    assert obstacles[0].name == "山"
    assert obstacles[0].mag_bearing_deg == 4.0
    assert obstacles[0].distance_m == 33630.0
    assert obstacles[0].elevation_m == 2484.0


def test_suspicious_table_success_is_repaired_from_text_fallback():
    table = _obs("082", 1.0, 36.0, elevation_m=None)
    table.name = "山体"
    table.raw_text = "RWY01/36L/36R"
    text = _obs("082", 151.0, 25664.0, elevation_m=2096.0)

    obstacles, failures = merge_table_and_text_results([table], [], [text])

    assert failures == []
    assert obstacles[0].name == "山体"
    assert obstacles[0].mag_bearing_deg == 151.0
    assert obstacles[0].distance_m == 25664.0
    assert obstacles[0].elevation_m == 2096.0
    assert obstacles[0].confidence_score == CONFIDENCE_OK


def test_missing_numeric_id_is_reported_as_failure():
    obstacles = [_obs(f"{i:03d}", 10.0, 1000.0) for i in range(1, 13) if i != 7]

    failures = append_missing_id_failures(obstacles, [])

    assert len(failures) == 1
    assert failures[0].airport_icao == "ZLLL"
    assert failures[0].source_page == 0
    assert "007" in failures[0].reason
    assert "007" in failures[0].raw_text


def test_missing_numeric_id_does_not_duplicate_existing_failure():
    obstacles = [_obs(f"{i:03d}", 10.0, 1000.0) for i in range(1, 13) if i != 7]
    existing = ParseFailure(
        airport_icao="ZLLL",
        source_page=3,
        raw_text="007",
        reason="编号 007：缺少方位/距离",
    )

    failures = append_missing_id_failures(obstacles, [existing])

    assert failures == [existing]


def test_missing_numeric_id_check_requires_enough_records():
    obstacles = [_obs(f"{i:03d}", 10.0, 1000.0) for i in range(1, 9) if i != 4]

    failures = append_missing_id_failures(obstacles, [])

    assert failures == []


def test_missing_numeric_id_check_ignores_non_numeric_ids():
    obstacles = [_obs(f"{i:03d}", 10.0, 1000.0) for i in range(1, 11)]
    obstacles.append(_obs("A-012", 10.0, 1000.0))

    failures = append_missing_id_failures(obstacles, [])

    assert failures == []


def test_pipeline_adds_missing_numeric_id_to_stats():
    obstacles = [_obs(f"{i:03d}", 10.0, 1000.0) for i in range(1, 13) if i != 7]

    result = _build_parse_result(
        source_file="test.pdf",
        file_hash="hash",
        obstacles=obstacles,
        failures=[],
        log_name="test.pdf",
        log_mode="test",
    )

    assert result.stats.total_success == 11
    assert result.stats.total_failed == 1
    assert result.stats.total_candidates == 12
    assert result.failures[0].reason.endswith("007")


_LANZHOU_PDF = Path(__file__).parent.parent / "examples" / "兰州_6-20.pdf"
_LANZHOU_XLSX = Path(__file__).parent.parent / "examples" / "兰州.xlsx"
_SKIP_LANZHOU = pytest.mark.skipif(
    not _LANZHOU_PDF.exists() or not _LANZHOU_XLSX.exists(),
    reason="Lanzhou example files not found",
)


@pytest.fixture(scope="module")
def lanzhou_result():
    if not _LANZHOU_PDF.exists():
        pytest.skip("Lanzhou PDF not found")
    return parse_file(_LANZHOU_PDF)


def _manual_rows() -> dict[str, dict[str, object]]:
    wb = load_workbook(_LANZHOU_XLSX, data_only=True)
    ws = wb[wb.sheetnames[0]]
    rows: dict[str, dict[str, object]] = {}
    for row in ws.iter_rows(min_row=3, values_only=True):
        if not row[0]:
            continue
        obstacle_id = str(row[0]).zfill(3)
        rows[obstacle_id] = {
            "name": row[1],
            "mag_bearing_deg": float(row[2]),
            "distance_m": float(row[3]),
            "elevation_m": float(row[4]),
            "modified": row[6],
        }
    return rows


@_SKIP_LANZHOU
def test_lanzhou_pdf_has_no_remaining_parse_failures(lanzhou_result):
    assert lanzhou_result.stats.total_success == 150
    assert lanzhou_result.stats.total_failed == 0
    assert lanzhou_result.failures == []


@_SKIP_LANZHOU
def test_lanzhou_manually_fixed_rows_match_core_fields(lanzhou_result):
    manual = _manual_rows()
    auto = {obs.obstacle_id: obs for obs in lanzhou_result.obstacles}
    modified_ids = [
        obstacle_id
        for obstacle_id, row in manual.items()
        if str(row["modified"]).strip() == "是"
    ]

    assert len(modified_ids) == 22
    for obstacle_id in modified_ids:
        obs = auto[obstacle_id]
        row = manual[obstacle_id]
        assert obs.mag_bearing_deg == row["mag_bearing_deg"]
        assert obs.distance_m == row["distance_m"]
        assert obs.elevation_m == row["elevation_m"]
        assert obs.confidence_score == CONFIDENCE_OK


@_SKIP_LANZHOU
def test_lanzhou_missing_coordinates_are_not_risk(lanzhou_result):
    no_coord = [
        obs for obs in lanzhou_result.obstacles
        if obs.latitude is None or obs.longitude is None
    ]

    assert no_coord
    assert all(obs.confidence_score == CONFIDENCE_OK for obs in no_coord)
